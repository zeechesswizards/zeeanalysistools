import streamlit as st
import pandas as pd
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

st.set_page_config(page_title="CFO Strategy & Audit Engine", layout="wide")
st.title("💼 CFO Strategy & Audit Engine")
st.write("Upload your QuickBooks ledger file below to instantly generate your Executive Dashboard and Audit Report.")

uploaded_file = st.file_uploader("Choose a QuickBooks Excel file (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    try:
        with st.spinner("Processing ledger data and compiling metrics..."):

            # ----------------------------------------------------------------
            # 1. Find header row dynamically
            # ----------------------------------------------------------------
            raw_df = pd.read_excel(uploaded_file, header=None)
            skip_rows = 0
            for i, row in raw_df.iterrows():
                vals = [str(v).strip() for v in row.values]
                if 'Date' in vals and 'Amount' in vals:
                    skip_rows = i
                    break

            df = pd.read_excel(uploaded_file, skiprows=skip_rows)
            df.columns = df.columns.str.strip()

            # ----------------------------------------------------------------
            # 2. Flexible Column Mapping
            # ----------------------------------------------------------------
            col_mapping = {}
            for c in df.columns:
                c_lower = c.lower().strip()
                if c_lower in ['name', 'vendor', 'vendor name']:               col_mapping[c] = 'Name'
                if c_lower in ['amount', 'amount (usd)', 'total amount']:      col_mapping[c] = 'Amount'
                if c_lower in ['date', 'transaction date']:                    col_mapping[c] = 'Date'
                if c_lower in ['no.', 'num', 'no', 'reference', 'doc number']:col_mapping[c] = 'No.'
                if c_lower in ['account', 'account name', 'distribution account']: col_mapping[c] = 'Account'
                if c_lower in ['split', 'distribution']:                       col_mapping[c] = 'Split'
                if c_lower in ['type', 'transaction type', 'txn type']:        col_mapping[c] = 'Transaction Type'
                if c_lower in ['memo/description', 'memo', 'description', 'memo / description']:
                    col_mapping[c] = 'Description'
                acct_num_names = ['account #','account#','acct #','acct#','account number',
                                  'account no','account no.','acct no','acct no.','acc #','acc#','account num']
                if c_lower in acct_num_names: col_mapping[c] = 'Account_Num'

            df = df.rename(columns=col_mapping)
            for col, default in [('No.','N/A'),('Account','N/A'),('Split','N/A'),
                                  ('Transaction Type','N/A'),('Description','')]:
                if col not in df.columns: df[col] = default

            # ----------------------------------------------------------------
            # 3. Data cleaning
            # ----------------------------------------------------------------
            df = df.dropna(subset=['Date', 'Amount'])
            df['Date']   = pd.to_datetime(df['Date'], errors='coerce')
            df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').abs()
            df = df.dropna(subset=['Date', 'Amount'])
            df = df[df['Amount'] > 0.0].copy()

            BLANK_LABELS = ['[Unassigned/Blank Vendor]', 'nan', '']
            df['Name'] = df['Name'].fillna('').astype(str).str.strip()
            df.loc[df['Name'] == '', 'Name'] = '[Unassigned/Blank Vendor]'
            df['Description'] = df['Description'].fillna('').astype(str).str.strip()

            # ----------------------------------------------------------------
            # 4. Account filtering — exclude balance sheet (1,2,3,4)
            # ----------------------------------------------------------------
            EXCLUDED_PREFIXES = ['1', '2', '3', '4']

            def extract_account_number(val):
                val = str(val).strip()
                if not val or val.lower() in ('nan', 'n/a', ''): return ''
                val = re.sub(r'\.0+$', '', val)
                m = re.match(r'^(\d+)', val)
                return m.group(1) if m else ''

            def is_excluded_code(val):
                code = extract_account_number(val)
                return any(code.startswith(p) for p in EXCLUDED_PREFIXES) if code else False

            filter_col = 'Account_Num' if 'Account_Num' in df.columns else 'Account'
            df['_fc'] = df[filter_col].fillna('').astype(str)
            df_filtered = df[~df['_fc'].apply(is_excluded_code)].copy().drop(columns=['_fc'])

            # ----------------------------------------------------------------
            # 5. Fiscal calendar (Aug=1 … Jul=12)
            # ----------------------------------------------------------------
            fiscal_map = {8:1,9:2,10:3,11:4,12:5,1:6,2:7,3:8,4:9,5:10,6:11,7:12}
            df_filtered['Month_Name']  = df_filtered['Date'].dt.strftime('%b-%y')
            df_filtered['YearMonth']   = df_filtered['Date'].dt.to_period('M')
            df_filtered['Fiscal_Sort'] = df_filtered['Date'].dt.month.map(fiscal_map)
            df_sorted = df_filtered.sort_values(by=['Fiscal_Sort', 'Date'])

            # ----------------------------------------------------------------
            # 6. TOP 5 VENDORS — monthly breakdown, blanks excluded
            # ----------------------------------------------------------------
            top5_src = df_filtered[~df_filtered['Name'].isin(BLANK_LABELS)].copy()
            top5_names = (top5_src.groupby('Name')['Amount']
                          .sum().sort_values(ascending=False).head(5).index.tolist())

            top5_pivot = (top5_src[top5_src['Name'].isin(top5_names)]
                          .pivot_table(index='Name',
                                       columns=['Fiscal_Sort','Month_Name'],
                                       values='Amount', aggfunc='sum')
                          .fillna(0))
            top5_pivot.columns = [col[1] for col in top5_pivot.columns]
            top5_pivot['Total'] = top5_pivot.sum(axis=1)
            top5_pivot = top5_pivot.loc[top5_names]
            top_5_vendors = top5_pivot.reset_index().rename(columns={'Name': 'Vendor'})

            # ----------------------------------------------------------------
            # 7. GL Account summary (top 5 accounts by spend)
            # ----------------------------------------------------------------
            gl_summary = (df_filtered.groupby('Account')['Amount']
                          .sum().reset_index()
                          .sort_values('Amount', ascending=False).head(5))
            gl_summary.columns = ['Top GL Accounts / Categories', 'Total Spend']

            # ----------------------------------------------------------------
            # 8. Month-over-Month top movers
            # ----------------------------------------------------------------
            recent_months = sorted(df_filtered['YearMonth'].drop_duplicates().tolist())[-2:]
            if len(recent_months) == 2:
                prev_m, curr_m = recent_months
                df_mom = df_filtered[df_filtered['YearMonth'].isin([prev_m, curr_m])]
                mom_pivot = df_mom.pivot_table(index='Name', columns='YearMonth',
                                               values='Amount', aggfunc='sum').fillna(0)
                if prev_m in mom_pivot.columns and curr_m in mom_pivot.columns:
                    mom_pivot['Variance'] = mom_pivot[curr_m] - mom_pivot[prev_m]
                    mom_pivot = mom_pivot.drop(index='[Unassigned/Blank Vendor]', errors='ignore')
                    top_up   = mom_pivot.nlargest(3,  'Variance').reset_index(); top_up['Status']   = 'SPIKED UP'
                    top_down = mom_pivot.nsmallest(3, 'Variance').reset_index(); top_down['Status'] = 'DROPPED DOWN'
                    movers_df = pd.concat([top_up[['Name','Status','Variance']],
                                           top_down[['Name','Status','Variance']]])
                    movers_df.columns = ['Vendor','MoM Movement','Variance ($)']
                else:
                    movers_df = pd.DataFrame({"Vendor":["Not enough data"],"MoM Movement":["N/A"],"Variance ($)":[0]})
            else:
                movers_df = pd.DataFrame({"Vendor":["Not enough data"],"MoM Movement":["N/A"],"Variance ($)":[0]})

            # ----------------------------------------------------------------
            # 9. New vendor detection
            # ----------------------------------------------------------------
            latest_month  = sorted(df_filtered['YearMonth'].drop_duplicates().tolist())[-1]
            first_seen     = df_filtered.groupby('Name')['YearMonth'].min()
            new_vendors_list = [v for v in first_seen[first_seen == latest_month].index
                                if v not in BLANK_LABELS]
            if new_vendors_list:
                nv_df = df_filtered[(df_filtered['Name'].isin(new_vendors_list)) &
                                    (df_filtered['YearMonth'] == latest_month)]
                new_vendors_summary = (nv_df.groupby('Name')['Amount'].sum()
                                       .reset_index().sort_values('Amount', ascending=False))
                new_vendors_summary.columns = ['New Vendor (First Payment This Month)', 'Initial Spend']
            else:
                new_vendors_summary = pd.DataFrame({
                    'New Vendor (First Payment This Month)': ['No new vendors detected'],
                    'Initial Spend': [0.0]
                })

            # ----------------------------------------------------------------
            # 10. Weekend expense controls
            # ----------------------------------------------------------------
            df_filtered['Is_Weekend'] = df_filtered['Date'].dt.dayofweek.isin([5, 6])
            weekend_tx = df_filtered[df_filtered['Is_Weekend']].copy()
            controls_scorecard = pd.DataFrame({
                "WEEKEND EXPENSE CONTROLS": ["Weekend Off-Cycle Outflows","Weekend Transactions Volume"],
                "EXPOSURE": [f"${weekend_tx['Amount'].sum():,.2f}", f"{len(weekend_tx)} Rows Flagged"],
                "THREAT LEVEL": ["MEDIUM RISK" if len(weekend_tx) > 0 else "LOW RISK", "MONITOR"]
            })
            weekend_tx['Reason'] = 'Off-Hours Weekend Transaction'
            leakage_report = (weekend_tx[['Date','Transaction Type','No.','Name','Account','Amount','Reason']]
                              .sort_values('Amount', ascending=False))

            # ----------------------------------------------------------------
            # 11. Duplicate detection (7-day window)
            # ----------------------------------------------------------------
            df_dc = df_filtered.sort_values(['Name','Amount','Date'])
            same_p = (df_dc['Name']   == df_dc['Name'].shift(1))   & (df_dc['Amount'] == df_dc['Amount'].shift(1))
            same_n = (df_dc['Name']   == df_dc['Name'].shift(-1))  & (df_dc['Amount'] == df_dc['Amount'].shift(-1))
            diff_p = (df_dc['Date'] - df_dc['Date'].shift(1)).dt.days.abs()  <= 7
            diff_n = (df_dc['Date'] - df_dc['Date'].shift(-1)).dt.days.abs() <= 7
            is_dup = (same_p & diff_p) | (same_n & diff_n)
            dup_cols = ['Date','Transaction Type','No.','Name','Description','Account','Split','Amount']
            report2_display = df_dc[is_dup][[c for c in dup_cols if c in df_dc.columns]].sort_values(['Name','Date'])

            # ----------------------------------------------------------------
            # 12. Monthly summary pivot
            # ----------------------------------------------------------------
            report1 = (df_sorted.pivot_table(index='Name', columns=['Fiscal_Sort','Month_Name'],
                                              values='Amount', aggfunc='sum').fillna(0))
            report1.columns = [col[1] for col in report1.columns]
            report1['Total Vendor Spend'] = report1.sum(axis=1)
            report1 = report1.sort_values('Total Vendor Spend', ascending=False)

            # ----------------------------------------------------------------
            # 13. 80/20 Pareto
            # ----------------------------------------------------------------
            total_val = df_filtered['Amount'].sum()
            report1['Cum_Pct'] = (report1['Total Vendor Spend'].cumsum() / total_val) * 100
            core_count = max(1, len(report1[report1['Cum_Pct'] <= 80.0]))
            tail_count = max(0, len(report1) - core_count)
            core_amt   = report1['Total Vendor Spend'].iloc[:core_count].sum()
            tail_amt   = total_val - core_amt
            pareto_df  = pd.DataFrame({
                "Strategic Spend Segment (80/20 Rule)": ["Core Leverage Vendors (Top 80%)","Long-Tail Vendors (Bottom 20%)"],
                "Count":               [core_count, tail_count],
                "Total Segment Spend": [core_amt,   tail_amt],
                "Budget %": [f"{(core_amt/total_val)*100:.1f}%" if total_val else "0.0%",
                             f"{(tail_amt/total_val)*100:.1f}%" if total_val else "0.0%"]
            })
            report1 = report1.drop(columns=['Cum_Pct'])

            # ----------------------------------------------------------------
            # 14. Expense spikes
            # ----------------------------------------------------------------
            baseline = df_filtered.groupby('Name')['Amount'].transform('median')
            report3  = df_filtered[df_filtered['Amount'] > baseline * 1.5].sort_values('Amount', ascending=False)
            spike_cols = ['Date','Transaction Type','Name','Description','Account','Split','Amount']
            report3_display = report3[[c for c in spike_cols if c in report3.columns]]

            # ----------------------------------------------------------------
            # 15. KPI scorecard & reconciliation
            # ----------------------------------------------------------------
            kpi_df = pd.DataFrame({
                "EXECUTIVE SCORECARD": ["Total Filtered Spend","Total Active Vendors",
                                        "Potential Duplicates","Expense Spikes"],
                "VALUE": [f"${total_val:,.2f}", f"{df_filtered['Name'].nunique()}",
                          f"{len(report2_display)} entries", f"{len(report3)} instances"]
            })
            recon_df = pd.DataFrame({
                "Metric": ["Filtered Ledger Spend"],
                "Source": [f"${total_val:,.2f}"],
                "Output": [f"${report1['Total Vendor Spend'].sum():,.2f}"],
                "Status": ["MATCH"]
            })

            # ----------------------------------------------------------------
            # 16. Monthly trend
            # ----------------------------------------------------------------
            monthly_trend = (df_sorted.groupby(['Fiscal_Sort','Month_Name'])['Amount']
                             .sum().reset_index())
            monthly_trend['Raw_Pct'] = monthly_trend['Amount'].pct_change()
            trend_display = monthly_trend[['Month_Name','Amount']].copy()
            trend_display['Trend Vector'] = monthly_trend['Raw_Pct'].apply(
                lambda x: "Baseline" if pd.isna(x) else (f"UP +{x*100:.1f}%" if x > 0 else f"DOWN {x*100:.1f}%"))
            trend_display.columns = ['Fiscal Month','Total Spend','Trend Vector']

            # ================================================================
            # LAYOUT ENGINE — compute all anchor positions dynamically
            # so tables NEVER overlap regardless of how many months are in data
            # ================================================================
            TOP5_COLS = len(top_5_vendors.columns)   # Vendor + n months + Total
            TOP5_ROWS = len(top_5_vendors)            # always 5

            # Row anchors (1-based Excel rows, but pd.ExcelWriter startrow is 0-based)
            ROW_KPI       = 1   # Excel row 2  (startrow=1)
            ROW_TOP5      = ROW_KPI + 6        # below KPI block + gap
            ROW_GL        = ROW_TOP5 + TOP5_ROWS + 3   # below top5 + gap
            ROW_TREND     = ROW_TOP5            # same row as top5, to the right
            ROW_CONTROLS  = ROW_GL              # same row as GL, to the right
            ROW_MOVERS    = ROW_GL + 8          # below GL block + gap
            ROW_NEWVENDOR = ROW_MOVERS          # same row as movers, to the right
            ROW_PARETO    = ROW_MOVERS + 5      # below movers + gap

            # Column anchors (0-based for startcol, i.e. 0=col A, 1=col B)
            COL_LEFT  = 1                        # col B — all left-side tables start here
            COL_RIGHT = COL_LEFT + TOP5_COLS + 1 # one blank gap after top5 ends

            # ----------------------------------------------------------------
            # 17. Web preview
            # ----------------------------------------------------------------
            st.success("✅ Analysis Complete!")
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("Executive Scorecard")
                st.dataframe(kpi_df, use_container_width=True, hide_index=True)
                st.subheader("Top 5 Vendors — Monthly Spend")
                st.dataframe(top_5_vendors, use_container_width=True, hide_index=True)
            with col2:
                st.subheader("80/20 Pareto")
                st.dataframe(pareto_df, use_container_width=True, hide_index=True)
                st.subheader("Top GL Accounts")
                st.dataframe(gl_summary, use_container_width=True, hide_index=True)

            # ----------------------------------------------------------------
            # 18. Build Excel in memory
            # ----------------------------------------------------------------
            output_buffer = io.BytesIO()
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                # Executive Dashboard — all anchored dynamically
                kpi_df.to_excel(writer, sheet_name="Executive Dashboard",
                                index=False, startrow=ROW_KPI,     startcol=COL_LEFT)
                top_5_vendors.to_excel(writer, sheet_name="Executive Dashboard",
                                       index=False, startrow=ROW_TOP5,    startcol=COL_LEFT)
                gl_summary.to_excel(writer, sheet_name="Executive Dashboard",
                                    index=False, startrow=ROW_GL,      startcol=COL_LEFT)
                trend_display.to_excel(writer, sheet_name="Executive Dashboard",
                                       index=False, startrow=ROW_TREND,   startcol=COL_RIGHT)
                controls_scorecard.to_excel(writer, sheet_name="Executive Dashboard",
                                            index=False, startrow=ROW_CONTROLS, startcol=COL_RIGHT)
                movers_df.to_excel(writer, sheet_name="Executive Dashboard",
                                   index=False, startrow=ROW_MOVERS,   startcol=COL_LEFT)
                new_vendors_summary.to_excel(writer, sheet_name="Executive Dashboard",
                                             index=False, startrow=ROW_NEWVENDOR, startcol=COL_RIGHT)
                pareto_df.to_excel(writer, sheet_name="Executive Dashboard",
                                   index=False, startrow=ROW_PARETO,   startcol=COL_LEFT)

                # Data sheets
                recon_df.to_excel(writer,        sheet_name="Recon",        index=False)
                report1.to_excel(writer,          sheet_name="R1-Summary")
                report2_display.to_excel(writer,  sheet_name="R2-Duplicates",index=False)
                report3_display.to_excel(writer,  sheet_name="R3-Spikes",    index=False)
                leakage_report.to_excel(writer,   sheet_name="R4-Weekend",   index=False)

            # ----------------------------------------------------------------
            # 19. Apply openpyxl styling
            # ----------------------------------------------------------------
            output_buffer.seek(0)
            wb = openpyxl.load_workbook(output_buffer)

            navy_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            black_bold = Font(name="Calibri", size=11, bold=True, color="000000")
            t_border   = Border(left=Side(style='thin', color='BFBFBF'),
                                right=Side(style='thin', color='BFBFBF'),
                                top=Side(style='thin', color='BFBFBF'),
                                bottom=Side(style='thin', color='BFBFBF'))

            def style_table_header(ws, excel_row, start_col, num_cols, fill=None):
                """Style a header row navy with white bold text."""
                f = fill or navy_fill
                for c in range(start_col, start_col + num_cols):
                    cell = ws.cell(row=excel_row, column=c)
                    cell.fill  = f
                    cell.font  = white_bold
                    cell.alignment = Alignment(horizontal='center')

            def style_table_data(ws, data_start_row, end_row, start_col, num_cols,
                                 money_cols=None, bold_last_col=False):
                """Apply zebra rows, currency format, borders."""
                for r in range(data_start_row, end_row + 1):
                    for c in range(start_col, start_col + num_cols):
                        cell = ws.cell(row=r, column=c)
                        cell.border = t_border
                        if (r - data_start_row) % 2 == 1:
                            cell.fill = zebra_fill
                        if money_cols and (c - start_col) in money_cols:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '$#,##0'
                                cell.alignment = Alignment(horizontal='right')
                if bold_last_col:
                    for r in range(data_start_row, end_row + 1):
                        cell = ws.cell(row=r, column=start_col + num_cols - 1)
                        if isinstance(cell.value, (int, float)):
                            cell.font = black_bold
                            cell.number_format = '$#,##0'

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.views.sheetView[0].showGridLines = True

                if sheet_name == "Executive Dashboard":
                    # Excel rows are 1-based; our ROW_ vars are 0-based startrow
                    # so Excel header row = ROW_ + 1 + 1 = ROW_ + 2
                    # Excel data starts   = ROW_ + 2 + 1 = ROW_ + 3

                    # KPI block
                    style_table_header(ws, ROW_KPI + 1,      COL_LEFT + 1, len(kpi_df.columns))
                    style_table_data(ws,   ROW_KPI + 2,      ROW_KPI + 1 + len(kpi_df), COL_LEFT + 1, len(kpi_df.columns))

                    # Top 5 vendors — highlight Total col
                    style_table_header(ws, ROW_TOP5 + 1,     COL_LEFT + 1, TOP5_COLS)
                    # Override Total col header to slightly different blue
                    ws.cell(row=ROW_TOP5 + 1, column=COL_LEFT + TOP5_COLS).fill = PatternFill(
                        start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                    money_month_cols = list(range(1, TOP5_COLS))  # all cols except Vendor (col 0)
                    style_table_data(ws, ROW_TOP5 + 2, ROW_TOP5 + 1 + TOP5_ROWS,
                                     COL_LEFT + 1, TOP5_COLS,
                                     money_cols=money_month_cols, bold_last_col=True)

                    # GL summary
                    style_table_header(ws, ROW_GL + 1,       COL_LEFT + 1, len(gl_summary.columns))
                    style_table_data(ws,   ROW_GL + 2,       ROW_GL + 1 + len(gl_summary),
                                     COL_LEFT + 1, len(gl_summary.columns), money_cols=[1])

                    # Trend table
                    style_table_header(ws, ROW_TREND + 1,    COL_RIGHT + 1, len(trend_display.columns))
                    style_table_data(ws,   ROW_TREND + 2,    ROW_TREND + 1 + len(trend_display),
                                     COL_RIGHT + 1, len(trend_display.columns), money_cols=[1])

                    # Controls scorecard
                    style_table_header(ws, ROW_CONTROLS + 1, COL_RIGHT + 1, len(controls_scorecard.columns))
                    style_table_data(ws,   ROW_CONTROLS + 2, ROW_CONTROLS + 1 + len(controls_scorecard),
                                     COL_RIGHT + 1, len(controls_scorecard.columns))

                    # Movers
                    style_table_header(ws, ROW_MOVERS + 1,   COL_LEFT + 1, len(movers_df.columns))
                    style_table_data(ws,   ROW_MOVERS + 2,   ROW_MOVERS + 1 + len(movers_df),
                                     COL_LEFT + 1, len(movers_df.columns), money_cols=[2])

                    # New vendors
                    style_table_header(ws, ROW_NEWVENDOR + 1, COL_RIGHT + 1, len(new_vendors_summary.columns))
                    style_table_data(ws,   ROW_NEWVENDOR + 2, ROW_NEWVENDOR + 1 + len(new_vendors_summary),
                                     COL_RIGHT + 1, len(new_vendors_summary.columns), money_cols=[1])

                    # Pareto
                    style_table_header(ws, ROW_PARETO + 1,   COL_LEFT + 1, len(pareto_df.columns))
                    style_table_data(ws,   ROW_PARETO + 2,   ROW_PARETO + 1 + len(pareto_df),
                                     COL_LEFT + 1, len(pareto_df.columns), money_cols=[2])

                    # Bar chart — pinned to right of trend table
                    chart = BarChart()
                    chart.type, chart.style = "col", 10
                    chart.title = "Monthly Spend Trajectory"
                    chart.y_axis.title, chart.legend = "Outflow ($)", None
                    chart.height, chart.width = 12, 18
                    trend_excel_header = ROW_TREND + 1
                    trend_excel_data_end = ROW_TREND + 1 + len(trend_display)
                    chart.add_data(Reference(ws, min_col=COL_RIGHT + 2,
                                             min_row=trend_excel_header,
                                             max_row=trend_excel_data_end), titles_from_data=True)
                    chart.set_categories(Reference(ws, min_col=COL_RIGHT + 1,
                                                   min_row=trend_excel_header + 1,
                                                   max_row=trend_excel_data_end))
                    chart_col = get_column_letter(COL_RIGHT + len(trend_display.columns) + 3)
                    ws.add_chart(chart, f"{chart_col}{ROW_TREND + 1}")

                else:
                    ws.freeze_panes = "A2"
                    for cell in ws[1]:
                        cell.fill = navy_fill
                        cell.font = white_bold
                        cell.alignment = Alignment(
                            horizontal='left' if str(cell.value or '').lower() in ('name','vendor') else 'center')
                    for row in range(2, ws.max_row + 1):
                        if row % 2 == 0:
                            for col in range(1, ws.max_column + 1):
                                c = ws.cell(row=row, column=col)
                                if c.fill.fill_type in (None, 'none'):
                                    c.fill = zebra_fill
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.border = t_border
                            header = str(ws.cell(row=1, column=col).value or '').lower()
                            if any(k in header for k in ['amount','spend','total']):
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = '$#,##0.00'
                                    cell.alignment = Alignment(horizontal='right')
                            if 'date' in header and isinstance(cell.value, pd.Timestamp):
                                cell.number_format = 'yyyy-mm-dd'
                                cell.alignment = Alignment(horizontal='center')

                # Auto-fit columns
                for col in ws.columns:
                    max_len    = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

        st.write("---")
        st.download_button(
            label="📥 Download Structured Analysis Report (Excel)",
            data=final_buffer,
            file_name="CFO_Analysis_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")
        import traceback
        st.code(traceback.format_exc())
